"""Step 2 — Header Reconstruction.

Handles complex Excel header structures: merged cells, hierarchical
(multi-level) headers, and flattening into a single header row.
"""

from __future__ import annotations

from typing import Any, Dict, List, Optional, Tuple

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

Region = Tuple[int, int, int, int]


def _get_merged_cell_map(ws: Worksheet) -> Dict[Tuple[int, int], Any]:
    """Build a mapping ``(row, col) -> value`` that propagates merged-cell
    values to every position covered by the merge.

    Both keys and values use **1-based** indexing (matching openpyxl).
    """
    merged_map: Dict[Tuple[int, int], Any] = {}
    for merged_range in ws.merged_cells.ranges:
        min_row = merged_range.min_row
        min_col = merged_range.min_col
        value = ws.cell(row=min_row, column=min_col).value
        for r in range(merged_range.min_row, merged_range.max_row + 1):
            for c in range(merged_range.min_col, merged_range.max_col + 1):
                merged_map[(r, c)] = value
    return merged_map


def _cell_value(
    ws: Worksheet,
    row: int,
    col: int,
    merged_map: Dict[Tuple[int, int], Any],
) -> Any:
    """Return the effective value of a cell, accounting for merges.

    Uses 1-based indexing.
    """
    if (row, col) in merged_map:
        return merged_map[(row, col)]
    return ws.cell(row=row, column=col).value


def detect_header_rows(
    ws: Worksheet,
    region: Region,
    merged_map: Dict[Tuple[int, int], Any],
    max_header_rows: int = 5,
) -> int:
    """Heuristically determine how many rows at the top of *region* are header rows.

    A row is considered a header row if **all** of its non-empty cells are
    strings (i.e., no numeric data).

    Parameters
    ----------
    ws : Worksheet
        The openpyxl worksheet.
    region : Region
        0-indexed ``(min_row, min_col, max_row, max_col)``.
    merged_map : dict
        Merged-cell value map (1-based keys).
    max_header_rows : int
        Upper bound on the number of header rows to detect.

    Returns
    -------
    int
        Number of header rows detected (may be 0).
    """
    min_r, min_c, max_r, max_c = region
    header_rows = 0
    for r_offset in range(min(max_header_rows, max_r - min_r + 1)):
        r = min_r + r_offset + 1  # 1-based
        all_str = True
        has_value = False
        for c in range(min_c + 1, max_c + 2):  # 1-based
            val = _cell_value(ws, r, c, merged_map)
            if val is not None:
                has_value = True
                if not isinstance(val, str):
                    all_str = False
                    break
        if has_value and all_str:
            header_rows += 1
        else:
            break
    return header_rows


def reconstruct_headers(
    ws: Worksheet,
    region: Region,
    n_header_rows: Optional[int] = None,
) -> List[str]:
    """Reconstruct flattened column headers for a table region.

    Supports hierarchical / multi-level headers created with merged cells.
    Headers are flattened by joining levels with ``"_"``.

    Parameters
    ----------
    ws : Worksheet
        The openpyxl worksheet.
    region : Region
        0-indexed ``(min_row, min_col, max_row, max_col)``.
    n_header_rows : int, optional
        Number of header rows.  Detected automatically if *None*.

    Returns
    -------
    list[str]
        One header string per column in the region.
    """
    merged_map = _get_merged_cell_map(ws)

    if n_header_rows is None:
        n_header_rows = detect_header_rows(ws, region, merged_map)
    if n_header_rows == 0:
        n_header_rows = 1  # assume at least one header row

    min_r, min_c, _, max_c = region
    n_cols = max_c - min_c + 1
    headers: List[str] = []

    for c_offset in range(n_cols):
        parts: List[str] = []
        for h_row in range(n_header_rows):
            r = min_r + h_row + 1  # 1-based
            c = min_c + c_offset + 1  # 1-based
            val = _cell_value(ws, r, c, merged_map)
            part = str(val).strip() if val is not None else ""
            if part and part not in parts:
                parts.append(part)
        header = "_".join(parts) if parts else f"col_{c_offset}"
        headers.append(header)

    return headers
