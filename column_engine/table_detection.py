"""Step 1 — Table Detection from raw Excel files.

Treats each Excel sheet as a 2D grid, detects non-empty cells, and groups
connected components using BFS.  Each connected block is a table candidate;
very small blocks are filtered out as noise.
"""

from __future__ import annotations

from collections import deque
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet


# Type alias for a rectangular region: (min_row, min_col, max_row, max_col)
Region = Tuple[int, int, int, int]


def _sheet_to_grid(ws: Worksheet) -> Tuple[List[List[Any]], int, int]:
    """Convert an openpyxl worksheet into a dense 2D list.

    Returns
    -------
    grid : list[list[Any]]
        ``grid[r][c]`` holds the cell value (or ``None``).
    n_rows, n_cols : int
        Dimensions of the grid.
    """
    n_rows = ws.max_row or 0
    n_cols = ws.max_column or 0
    grid: List[List[Any]] = [
        [None] * n_cols for _ in range(n_rows)
    ]
    for row in ws.iter_rows(min_row=1, max_row=n_rows,
                            min_col=1, max_col=n_cols):
        for cell in row:
            grid[cell.row - 1][cell.column - 1] = cell.value
    return grid, n_rows, n_cols


def _bfs(
    grid: List[List[Any]],
    start: Tuple[int, int],
    visited: set,
    n_rows: int,
    n_cols: int,
) -> List[Tuple[int, int]]:
    """BFS over non-empty cells starting from *start*.

    Returns a list of ``(row, col)`` positions belonging to the connected
    component.
    """
    queue: deque[Tuple[int, int]] = deque([start])
    visited.add(start)
    component: List[Tuple[int, int]] = []
    while queue:
        r, c = queue.popleft()
        component.append((r, c))
        for dr, dc in ((-1, 0), (1, 0), (0, -1), (0, 1)):
            nr, nc = r + dr, c + dc
            if 0 <= nr < n_rows and 0 <= nc < n_cols and (nr, nc) not in visited:
                if grid[nr][nc] is not None:
                    visited.add((nr, nc))
                    queue.append((nr, nc))
    return component


def _component_to_region(component: List[Tuple[int, int]]) -> Region:
    """Return the bounding-box region of a connected component."""
    rows = [r for r, _ in component]
    cols = [c for _, c in component]
    return (min(rows), min(cols), max(rows), max(cols))


def detect_tables(
    ws: Worksheet,
    min_cells: int = 4,
) -> List[Region]:
    """Detect table regions in an Excel worksheet.

    Parameters
    ----------
    ws : Worksheet
        An openpyxl worksheet.
    min_cells : int
        Minimum number of non-empty cells for a connected component to be
        kept as a table candidate (filters noise).

    Returns
    -------
    list[Region]
        List of ``(min_row, min_col, max_row, max_col)`` regions (0-indexed).
    """
    grid, n_rows, n_cols = _sheet_to_grid(ws)
    visited: set[Tuple[int, int]] = set()
    regions: List[Region] = []

    for r in range(n_rows):
        for c in range(n_cols):
            if grid[r][c] is not None and (r, c) not in visited:
                component = _bfs(grid, (r, c), visited, n_rows, n_cols)
                if len(component) >= min_cells:
                    regions.append(_component_to_region(component))

    return regions


def detect_tables_from_file(
    filepath: str,
    sheet_name: Optional[str] = None,
    min_cells: int = 4,
) -> Dict[str, List[Region]]:
    """Detect tables across all (or a specified) sheet of an Excel file.

    Parameters
    ----------
    filepath : str
        Path to the ``.xlsx`` file.
    sheet_name : str, optional
        Process only this sheet.  If *None*, processes every sheet.
    min_cells : int
        Forwarded to :func:`detect_tables`.

    Returns
    -------
    dict[str, list[Region]]
        Mapping ``sheet_name -> list_of_regions``.
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheets = [sheet_name] if sheet_name else wb.sheetnames
    result: Dict[str, List[Region]] = {}
    for name in sheets:
        ws = wb[name]
        result[name] = detect_tables(ws, min_cells=min_cells)
    wb.close()
    return result
