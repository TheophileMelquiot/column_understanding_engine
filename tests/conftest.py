"""Shared test fixtures: Excel workbook generators for unit tests."""

from __future__ import annotations

import os
import tempfile
from typing import Optional

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet


def create_simple_workbook(path: str) -> str:
    """Create a simple single-table workbook and return *path*."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    # Header
    ws["A1"] = "Name"
    ws["B1"] = "Age"
    ws["C1"] = "Email"
    # Data rows
    data = [
        ("Alice", 30, "alice@example.com"),
        ("Bob", 25, "bob@example.com"),
        ("Charlie", 35, "charlie@example.com"),
    ]
    for i, (name, age, email) in enumerate(data, start=2):
        ws.cell(row=i, column=1, value=name)
        ws.cell(row=i, column=2, value=age)
        ws.cell(row=i, column=3, value=email)
    wb.save(path)
    wb.close()
    return path


def create_multi_table_workbook(path: str) -> str:
    """Create a workbook with two separate tables on one sheet."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    # Table 1 (rows 1-4, cols A-C)
    ws["A1"] = "Product"
    ws["B1"] = "Price"
    ws["C1"] = "Qty"
    for i, (prod, price, qty) in enumerate(
        [("Widget", 9.99, 100), ("Gadget", 19.99, 50), ("Doohickey", 4.99, 200)],
        start=2,
    ):
        ws.cell(row=i, column=1, value=prod)
        ws.cell(row=i, column=2, value=price)
        ws.cell(row=i, column=3, value=qty)

    # Table 2 (rows 7-10, cols A-B)  — gap at rows 5-6
    ws["A7"] = "City"
    ws["B7"] = "Population"
    for i, (city, pop) in enumerate(
        [("Paris", 2161000), ("London", 8982000), ("Berlin", 3748000)],
        start=8,
    ):
        ws.cell(row=i, column=1, value=city)
        ws.cell(row=i, column=2, value=pop)

    wb.save(path)
    wb.close()
    return path


def create_merged_header_workbook(path: str) -> str:
    """Create a workbook with merged hierarchical headers."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Merged header: "Sales" spans B1:C1
    ws["A1"] = ""
    ws["B1"] = "Sales"
    ws.merge_cells("B1:C1")

    # Sub-headers
    ws["A2"] = "Year"
    ws["B2"] = "Q1"
    ws["C2"] = "Q2"

    # Data
    ws["A3"] = 2023
    ws["B3"] = 1000
    ws["C3"] = 1500
    ws["A4"] = 2024
    ws["B4"] = 1200
    ws["C4"] = 1700

    wb.save(path)
    wb.close()
    return path


def create_noisy_workbook(path: str) -> str:
    """Create a workbook with noisy/sparse data around a valid table."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Isolated noise cell
    ws["F1"] = "noise"

    # Main table at rows 3-6, cols A-C
    ws["A3"] = "ID"
    ws["B3"] = "Value"
    ws["C3"] = "Date"
    ws["A4"] = 1
    ws["B4"] = 42.5
    ws["C4"] = "2024-01-15"
    ws["A5"] = 2
    ws["B5"] = 38.1
    ws["C5"] = "2024-02-20"
    ws["A6"] = 3
    ws["B6"] = 55.0
    ws["C6"] = "2024-03-10"

    wb.save(path)
    wb.close()
    return path
