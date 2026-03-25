"""Tests for Step 3 — Data Cleaning."""

from __future__ import annotations

import openpyxl
import pandas as pd
import pytest

from column_engine.data_cleaning import clean_table, clean_tables_from_sheet
from column_engine.table_detection import detect_tables
from tests.conftest import create_simple_workbook, create_multi_table_workbook


@pytest.fixture()
def simple_xlsx(tmp_path):
    return create_simple_workbook(str(tmp_path / "simple.xlsx"))


@pytest.fixture()
def multi_xlsx(tmp_path):
    return create_multi_table_workbook(str(tmp_path / "multi.xlsx"))


class TestCleanTable:
    def test_simple_clean(self, simple_xlsx):
        wb = openpyxl.load_workbook(simple_xlsx)
        ws = wb.active
        regions = detect_tables(ws, min_cells=2)
        assert len(regions) == 1
        df = clean_table(ws, regions[0])
        assert isinstance(df, pd.DataFrame)
        assert len(df) == 3  # 3 data rows
        assert "Name" in df.columns
        assert "Age" in df.columns
        assert "Email" in df.columns
        wb.close()

    def test_values_preserved(self, simple_xlsx):
        wb = openpyxl.load_workbook(simple_xlsx)
        ws = wb.active
        regions = detect_tables(ws, min_cells=2)
        df = clean_table(ws, regions[0])
        assert df["Name"].tolist() == ["Alice", "Bob", "Charlie"]
        wb.close()


class TestCleanTablesFromSheet:
    def test_multi_table(self, multi_xlsx):
        wb = openpyxl.load_workbook(multi_xlsx)
        ws = wb.active
        regions = detect_tables(ws, min_cells=2)
        tables = clean_tables_from_sheet(ws, regions)
        assert len(tables) == 2
        for key, df in tables.items():
            assert isinstance(df, pd.DataFrame)
            assert len(df) > 0
        wb.close()
