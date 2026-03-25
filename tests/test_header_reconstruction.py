"""Tests for Step 2 — Header Reconstruction."""

from __future__ import annotations

import openpyxl
import pytest

from column_engine.header_reconstruction import (
    _get_merged_cell_map,
    reconstruct_headers,
)
from tests.conftest import create_merged_header_workbook, create_simple_workbook


@pytest.fixture()
def merged_xlsx(tmp_path):
    return create_merged_header_workbook(str(tmp_path / "merged.xlsx"))


@pytest.fixture()
def simple_xlsx(tmp_path):
    return create_simple_workbook(str(tmp_path / "simple.xlsx"))


class TestMergedCellMap:
    def test_merged_propagation(self, merged_xlsx):
        wb = openpyxl.load_workbook(merged_xlsx)
        ws = wb.active
        merged_map = _get_merged_cell_map(ws)
        # B1:C1 merged with value "Sales"
        assert merged_map[(1, 2)] == "Sales"
        assert merged_map[(1, 3)] == "Sales"
        wb.close()


class TestReconstructHeaders:
    def test_simple_headers(self, simple_xlsx):
        wb = openpyxl.load_workbook(simple_xlsx)
        ws = wb.active
        # Region: rows 0-3, cols 0-2 (0-indexed)
        headers = reconstruct_headers(ws, (0, 0, 3, 2), n_header_rows=1)
        assert headers == ["Name", "Age", "Email"]
        wb.close()

    def test_hierarchical_headers(self, merged_xlsx):
        wb = openpyxl.load_workbook(merged_xlsx)
        ws = wb.active
        # Region: rows 0-3, cols 0-2 (0-indexed)
        headers = reconstruct_headers(ws, (0, 0, 3, 2), n_header_rows=2)
        # Column 0: row1="" row2="Year" → "Year"
        # Column 1: row1="Sales" row2="Q1" → "Sales_Q1"
        # Column 2: row1="Sales" row2="Q2" → "Sales_Q2"
        assert headers[0] == "Year"
        assert headers[1] == "Sales_Q1"
        assert headers[2] == "Sales_Q2"
        wb.close()

    def test_auto_detect_headers(self, simple_xlsx):
        wb = openpyxl.load_workbook(simple_xlsx)
        ws = wb.active
        headers = reconstruct_headers(ws, (0, 0, 3, 2))
        assert "Name" in headers
        wb.close()
