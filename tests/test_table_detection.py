"""Tests for Step 1 — Table Detection."""

from __future__ import annotations

import os
import tempfile

import openpyxl
import pytest

from column_engine.table_detection import detect_tables, detect_tables_from_file
from tests.conftest import (
    create_simple_workbook,
    create_multi_table_workbook,
    create_noisy_workbook,
)


@pytest.fixture()
def simple_xlsx(tmp_path):
    return create_simple_workbook(str(tmp_path / "simple.xlsx"))


@pytest.fixture()
def multi_xlsx(tmp_path):
    return create_multi_table_workbook(str(tmp_path / "multi.xlsx"))


@pytest.fixture()
def noisy_xlsx(tmp_path):
    return create_noisy_workbook(str(tmp_path / "noisy.xlsx"))


class TestDetectTables:
    def test_single_table(self, simple_xlsx):
        wb = openpyxl.load_workbook(simple_xlsx)
        ws = wb.active
        regions = detect_tables(ws, min_cells=2)
        assert len(regions) == 1
        min_r, min_c, max_r, max_c = regions[0]
        # Should span rows 0-3, cols 0-2 (0-indexed)
        assert min_r == 0
        assert min_c == 0
        assert max_r == 3
        assert max_c == 2
        wb.close()

    def test_multi_table(self, multi_xlsx):
        wb = openpyxl.load_workbook(multi_xlsx)
        ws = wb.active
        regions = detect_tables(ws, min_cells=2)
        assert len(regions) == 2
        wb.close()

    def test_noise_filtering(self, noisy_xlsx):
        wb = openpyxl.load_workbook(noisy_xlsx)
        ws = wb.active
        # With min_cells=4, the single noise cell should be filtered
        regions = detect_tables(ws, min_cells=4)
        assert len(regions) == 1
        wb.close()

    def test_detect_from_file(self, simple_xlsx):
        result = detect_tables_from_file(simple_xlsx)
        assert "Sheet1" in result
        assert len(result["Sheet1"]) == 1


class TestEmptySheet:
    def test_empty_sheet(self, tmp_path):
        path = str(tmp_path / "empty.xlsx")
        wb = openpyxl.Workbook()
        wb.save(path)
        wb.close()

        wb = openpyxl.load_workbook(path)
        ws = wb.active
        regions = detect_tables(ws, min_cells=1)
        assert regions == []
        wb.close()
